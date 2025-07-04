import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, filedialog
import pandas as pd
from datetime import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from questions_data import FORMS_DATA
import threading
import queue

class FormWizardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Medical History Form")
        self.root.geometry("800x700")
        self.filepath = None

        # This method will prompt the user and set self.filepath
        self._setup_filepath_on_startup()

        # If user cancels file selection, close the app
        if not self.filepath:
            self.root.destroy()
            return

        self.form_keys = list(FORMS_DATA.keys())
        self.current_page = 0
        self.patient_data = {} # Will use a unique key: (page_index, question_label)
        self.pages = {}
        self.widget_map = {}
        self.populated_pages = set()
        self.submit_queue = queue.Queue()

        # Main container for the pages
        self.page_container = tk.Frame(root)
        self.page_container.pack(side="top", fill="both", expand=True)
        self.page_container.grid_rowconfigure(0, weight=1)
        self.page_container.grid_columnconfigure(0, weight=1)

        self._create_pages()
        self._create_navigation()
        self.show_page(0)

    def _setup_filepath_on_startup(self):
        # Simple dialog, as a full custom Toplevel can be complex to manage state
        answer = messagebox.askquestion(
            "Setup Response File",
            "Do you want to open an existing response file?\n\n"
            "(Choose 'Yes' to append to an existing file, 'No' to create a new one.)",
            icon='question',
            type='yesnocancel'
        )

        if answer == 'yes':
            self.filepath = filedialog.askopenfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Open Existing Response File"
            )
        elif answer == 'no':
            self.filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Create New Response File",
                initialfile="patient_responses.xlsx"
            )
        else: # 'cancel' or window closed
            self.filepath = None

    def _bind_scroll(self, widget, canvas):
        """Binds mouse scroll events to a widget to scroll a canvas."""
        # For Windows/macOS
        widget.bind("<MouseWheel>", lambda e, c=canvas: c.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        # For Linux
        widget.bind("<Button-4>", lambda e, c=canvas: c.yview_scroll(-1, "units"))
        widget.bind("<Button-5>", lambda e, c=canvas: c.yview_scroll(1, "units"))

    def _bind_recursive(self, widget, canvas):
        """Binds mouse scroll events recursively to a widget and all its children."""
        self._bind_scroll(widget, canvas)
        for child in widget.winfo_children():
            self._bind_recursive(child, canvas)

    def _create_pages(self):
        for i, form_key in enumerate(self.form_keys):
            form_data = FORMS_DATA[form_key]
            frame = tk.Frame(self.page_container, padx=10, pady=10)
            self.pages[i] = {"frame": frame}
            frame.grid(row=0, column=0, sticky="nsew")

            title = ttk.Label(frame, text=form_data["title"], font=("Arial", 16, "bold"))
            title.pack(pady=(0, 20))

            # Canvas and Scrollbar for scrollable content
            canvas = tk.Canvas(frame)
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e, c=canvas: c.configure(scrollregion=c.bbox("all"))
            )
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Store metadata for lazy loading
            self.pages[i]["canvas"] = canvas
            self.pages[i]["scrollable_frame"] = scrollable_frame

    def _populate_page(self, parent, questions, page_index):
        self.widget_map[parent] = {}
        for question in questions:
            row_frame = tk.Frame(parent)
            row_frame.pack(fill="x", pady=2)
            
            if isinstance(question, dict) and question.get("type") == "dropdown":
                q_label = question["label"]
                unique_key = (page_index, q_label)

                display_options, value_map = [], {}
                for option in question["options"]:
                    parts = option.split('=', 1)
                    value = parts[0].strip()
                    display_text = f"{value}: {parts[1].strip()}" if len(parts) > 1 else value
                    display_options.append(display_text)
                    value_map[display_text] = value

                actual_value_var = self.patient_data.get(unique_key)
                if not isinstance(actual_value_var, tk.StringVar):
                    actual_value_var = tk.StringVar()
                    self.patient_data[unique_key] = actual_value_var

                display_var = tk.StringVar()

                label = ttk.Label(row_frame, text=q_label, width=40)
                label.pack(side="left", padx=5)

                combobox = ttk.Combobox(row_frame, textvariable=display_var, values=display_options, state="readonly", width=37)
                combobox.pack(side="left", fill="x", expand=True, padx=5)

                def on_select(event, var=actual_value_var, vmap=value_map, d_var=display_var):
                    actual_value = vmap.get(d_var.get(), "")
                    var.set(actual_value)
                
                combobox.bind("<<ComboboxSelected>>", on_select)

                saved_value = actual_value_var.get()
                if saved_value:
                    for text, val in value_map.items():
                        if val == saved_value:
                            display_var.set(text)
                            break
            
            elif isinstance(question, dict) and question.get("type") in ["yesno", "yesno_conditional"]:
                q_label = question["label"]
                unique_key = (page_index, q_label)
                var = tk.StringVar(value="No")
                
                # Use existing var if it's there (from loading data)
                if unique_key in self.patient_data and isinstance(self.patient_data.get(unique_key), tk.StringVar):
                    var = self.patient_data[unique_key]
                else:
                    self.patient_data[unique_key] = var

                label = ttk.Label(row_frame, text=q_label, width=40)
                label.pack(side="left", padx=5)

                yes_rb = ttk.Radiobutton(row_frame, text="Yes", variable=var, value="Yes")
                no_rb = ttk.Radiobutton(row_frame, text="No", variable=var, value="No")
                yes_rb.pack(side="left")
                no_rb.pack(side="left")
                
                if question.get("type") == "yesno_conditional":
                    self.widget_map[parent][q_label] = {
                        "hides": question["hides"], 
                        "widgets_to_hide": []
                    }
                    var.trace_add("write", lambda *args, q=question, p=parent, key=unique_key: self._toggle_visibility(q, p, key))

            else: # Standard question (text entry)
                q_label = question
                unique_key = (page_index, q_label)
                var = tk.StringVar()
                self.patient_data[unique_key] = var

                label = ttk.Label(row_frame, text=q_label, width=40)
                entry = ttk.Entry(row_frame, textvariable=var, width=50)
                label.pack(side="left", padx=5)
                entry.pack(side="left", padx=5, fill="x", expand=True)
                
                # If this widget can be hidden, add it to the map
                for conditional_q, data in self.widget_map.get(parent, {}).items():
                    if q_label in data["hides"]:
                        data["widgets_to_hide"].append(row_frame)

    def _toggle_visibility(self, question_data, parent, key):
        var = self.patient_data[key]
        is_visible = var.get() == "Yes"
        
        widgets_to_hide = self.widget_map[parent][question_data["label"]]["widgets_to_hide"]
        for widget_row in widgets_to_hide:
            if is_visible:
                widget_row.pack(fill="x", pady=2)
            else:
                widget_row.pack_forget()

    def _create_navigation(self):
        nav_frame = tk.Frame(self.root, pady=10)
        nav_frame.pack(side="bottom", fill="x")

        self.back_button = ttk.Button(nav_frame, text="Back", command=self.prev_page, state="disabled")
        self.next_button = ttk.Button(nav_frame, text="Next", command=self.next_page)
        self.submit_button = ttk.Button(nav_frame, text="Submit", command=self.submit, state="disabled")

        self.back_button.pack(side="left", padx=20)
        self.submit_button.pack(side="right", padx=20)
        self.next_button.pack(side="right")

    def show_page(self, page_number):
        if page_number in self.pages:
            self.current_page = page_number
            page_meta = self.pages[page_number]
            frame = page_meta["frame"]

            # Lazy load page content
            if page_number not in self.populated_pages:
                form_key = self.form_keys[page_number]
                form_data = FORMS_DATA[form_key]
                scrollable_frame = page_meta["scrollable_frame"]
                canvas = page_meta["canvas"]

                self._populate_page(scrollable_frame, form_data["questions"], page_number)
                self._bind_recursive(scrollable_frame, canvas)

                padding_frame = ttk.Frame(scrollable_frame, height=20)
                padding_frame.pack()
                self._bind_scroll(padding_frame, canvas)

                self.populated_pages.add(page_number)

            frame.tkraise()

            self.back_button["state"] = "normal" if self.current_page > 0 else "disabled"
            self.next_button["state"] = "normal" if self.current_page < len(self.pages) - 1 else "disabled"
            self.submit_button["state"] = "normal" if self.current_page == len(self.pages) - 1 else "disabled"

    def next_page(self):
        next_page_index = self.current_page + 1
        
        # Handle preliminary questions for skipping forms
        if next_page_index < len(self.form_keys):
            form_key = self.form_keys[next_page_index]
            form_data = FORMS_DATA[form_key]
            
            if "preliminary_question" in form_data:
                proceed = messagebox.askyesno(
                    form_data["title"],
                    form_data["preliminary_question"],
                    parent=self.root
                )
                if not proceed:
                    # Mark all questions in this form to be skipped by storing the value directly
                    fill_value = form_data.get("fill_value", "N/A")
                    for q in form_data["questions"]:
                        base_label = q['label'] if isinstance(q, dict) else q
                        unique_key = (next_page_index, base_label)
                        self.patient_data[unique_key] = fill_value
                    
                    # Try to skip to the page after this one
                    self.current_page += 1
                    self.next_page()
                    return

        self.show_page(next_page_index)

    def prev_page(self):
        self.show_page(self.current_page - 1)

    def submit(self):
        """Initiates the data submission process in a background thread."""
        self.submit_button.config(state="disabled", text="Saving...")
        self.back_button.config(state="disabled")
        self.next_button.config(state="disabled")

        # Offload the actual work to a background thread
        threading.Thread(target=self._submit_worker, daemon=True).start()
        
        # Start polling the queue for a result from the worker thread
        self.root.after(100, self._check_submit_queue)

    def _check_submit_queue(self):
        """Checks if the background submission task is done and handles the result."""
        try:
            result = self.submit_queue.get_nowait()
            # Re-enable buttons and reset text regardless of outcome
            self.submit_button.config(text="Submit")
            self.show_page(self.current_page) # This will correctly set button states

            if result['status'] == 'success':
                messagebox.showinfo(
                    "Success",
                    f"Data saved successfully to\n{result['filepath']}\n\nThe form has been reset for the next patient."
                )
                self._reset_form()
            elif result['status'] == 'error':
                messagebox.showerror("Error", f"An error occurred while saving:\n{result['error']}")

        except queue.Empty:
            # If the queue is empty, it means the worker is still busy.
            # Check again after a short delay.
            self.root.after(100, self._check_submit_queue)

    def _submit_worker(self):
        """The actual data processing and file writing task. Runs in a background thread."""
        if not self.filepath:
            self.submit_queue.put({"status": "error", "error": "No file path specified. Please restart."})
            return

        try:
            final_data = {"Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
            
            # This part is safe to run in a background thread as it only reads from thread-safe tkinter variables
            for i, form_key in enumerate(self.form_keys):
                form_info = FORMS_DATA[form_key]
                for question in form_info["questions"]:
                    if isinstance(question, dict) and question.get("type") == "yesno_conditional":
                        base_label = question['label']
                        unique_data_key = (i, base_label)
                        value_var = self.patient_data.get(unique_data_key)
                        
                        if value_var and isinstance(value_var, tk.StringVar) and value_var.get() == "No":
                            fill_value = question.get("fill_value", "N/A")
                            for hidden_q_label in question['hides']:
                                hidden_page_index = self._find_question_page(hidden_q_label)
                                if hidden_page_index != -1:
                                    hidden_key = (hidden_page_index, hidden_q_label)
                                    if hidden_key in self.patient_data and isinstance(self.patient_data[hidden_key], tk.StringVar):
                                         self.patient_data[hidden_key].set(fill_value)
                                    else:
                                         self.patient_data[hidden_key] = fill_value

            column_counts = {}
            for i, form_key in enumerate(self.form_keys):
                form_info = FORMS_DATA[form_key]
                for question in form_info["questions"]:
                    base_label = question['label'] if isinstance(question, dict) else question
                    
                    count = column_counts.get(base_label, 0) + 1
                    column_counts[base_label] = count
                    unique_col_name = f"{base_label}_{count}" if count > 1 else base_label

                    unique_data_key = (i, base_label)
                    value = self.patient_data.get(unique_data_key)

                    if isinstance(value, str):
                        final_data[unique_col_name] = value
                    elif isinstance(value, tk.StringVar):
                        final_data[unique_col_name] = value.get()
                    else:
                        final_data[unique_col_name] = "" # Failsafe
            
            new_entry_df = pd.DataFrame([final_data])
            
            if os.path.exists(self.filepath):
                existing_df = pd.read_excel(self.filepath)
                
                final_ordered_columns = existing_df.columns.tolist()
                for col in new_entry_df.columns:
                    if col not in final_ordered_columns:
                        final_ordered_columns.append(col)

                existing_df = existing_df.reindex(columns=final_ordered_columns)
                new_entry_df = new_entry_df.reindex(columns=final_ordered_columns)
                df_to_save = pd.concat([existing_df, new_entry_df], ignore_index=True)
            else:
                df_to_save = new_entry_df

            with pd.ExcelWriter(self.filepath, engine='openpyxl') as writer:
                df_to_save.to_excel(writer, index=False, sheet_name='Responses')
                
                worksheet = writer.sheets['Responses']
                
                section_fills = [
                    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
                    PatternFill(start_color="F2B4D0", end_color="F2B4D0", fill_type="solid"),
                    PatternFill(start_color="D9D9F3", end_color="D9D9F3", fill_type="solid"),
                    PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid"),
                    PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid"),
                    PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid"),
                    PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
                    PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
                    PatternFill(start_color="BDB76B", end_color="BDB76B", fill_type="solid"),
                    PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                ]

                section_color_map = {}
                color_index = 0
                for form_key in self.form_keys:
                    fill = section_fills[color_index % len(section_fills)]
                    for q in FORMS_DATA[form_key]["questions"]:
                        q_label = q["label"] if isinstance(q, dict) else q
                        section_color_map[q_label] = fill
                    color_index += 1

                for i, column_name in enumerate(df_to_save.columns):
                    column_letter = get_column_letter(i + 1)
                    
                    base_label = column_name.rsplit('_', 1)[0] if '_' in column_name and column_name.rsplit('_', 1)[1].isdigit() else column_name
                    if base_label in section_color_map:
                        worksheet.cell(row=1, column=i + 1).fill = section_color_map[base_label]

                    max_length = len(str(column_name))
                    for cell in worksheet[column_letter]:
                        if cell.row == 1:
                            continue
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            self.submit_queue.put({"status": "success", "filepath": self.filepath})

        except Exception as e:
            self.submit_queue.put({"status": "error", "error": e})

    def _reset_form(self):
        """Resets the form for a new patient entry. Should only be called from the main thread."""
        # Reset state
        self.current_page = 0
        self.patient_data = {}
        self.widget_map = {}
        self.populated_pages.clear()

        # Destroy all existing page frames' content
        for page_meta in self.pages.values():
            for widget in page_meta["frame"].winfo_children():
                widget.destroy()
        self.pages = {}

        # Re-create the pages from scratch
        self._create_pages()
        self.show_page(0)

    def _find_question_page(self, q_label):
        for i, form_key in enumerate(self.form_keys):
            for q in FORMS_DATA[form_key]["questions"]:
                label = q['label'] if isinstance(q, dict) else q
                if label == q_label:
                    return i
        return -1

if __name__ == "__main__":
    root = tk.Tk()
    app = FormWizardApp(root)
    root.mainloop() 